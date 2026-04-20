"""Build budget_full_{YYYY-MM}.xlsx from transactions.json + recurring.json.

The output matches the structure of the user's reference
`~/Downloads/budget_full_mar2026.xlsx`: a single "Monthly Budget" sheet with
sections for INCOME, FIXED OBLIGATIONS, MSI, SUBSCRIPTIONS, OFF-CARD,
VARIABLE, BUSINESS REIMBURSABLE, FEES, SUMMARY, and MSI ROLL-OFF TRACKER.

Amounts are MXN. %-income cells are Excel formulas against B6 (net income) so
the sheet remains editable.
"""

from __future__ import annotations

import datetime as dt
import os
from collections import defaultdict
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from utils.data_loader import load_transactions, load_recurring


HEADER_FILL = PatternFill("solid", fgColor="1F3864")  # dark blue
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")  # pale blue
SECTION_FONT = Font(bold=True)
SUBTOTAL_FONT = Font(bold=True)
THIN = Side(style="thin", color="B4C7E7")
ALL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MXN_FMT = "#,##0.00"
PCT_FMT = "0.0%"


def _parse_ym(month_str: str) -> tuple[int, int]:
    y, m = month_str.split("-")
    return int(y), int(m)


def _add_months(year: int, month: int, delta: int) -> tuple[int, int]:
    m0 = (month - 1) + delta
    ny, nm = divmod(m0, 12)
    return year + ny, nm + 1


def _month_label(year: int, month: int) -> str:
    return dt.date(year, month, 1).strftime("%B %Y")


def _sum_by_subcategory(
    transactions: list[dict], section: str, close_month: str
) -> dict[str, float]:
    totals: dict[str, float] = defaultdict(float)
    for t in transactions:
        if t.get("close_month") != close_month:
            continue
        if t.get("section") != section:
            continue
        sub = (t.get("subcategory") or "").strip() or "Other"
        amt = t.get("amount_mxn")
        if amt is None:
            amt = t.get("amount_native", 0)
        totals[sub] += float(amt or 0)
    return dict(totals)


def _msi_active_for_month(msi_rows: list[dict], close_month: str) -> list[dict]:
    """Return MSI rows that are still being paid in `close_month`."""
    y, m = _parse_ym(close_month)
    out = []
    for r in msi_rows:
        try:
            sy, sm = _parse_ym(r.get("start", "1970-01"))
            months_total = int(r.get("months_total", 0))
        except Exception:
            continue
        ey, em = _add_months(sy, sm, months_total - 1)
        # active if close_month is between start and end (inclusive)
        if (sy, sm) <= (y, m) <= (ey, em):
            # compute current installment number
            cur = (y - sy) * 12 + (m - sm) + 1
            out.append({**r, "current_installment": cur, "end_year": ey, "end_month": em})
    return out


def _msi_ending_soon(msi_rows: list[dict], close_month: str, horizon_months: int = 3) -> list[dict]:
    """MSI rows ending in `close_month` + next `horizon_months`."""
    y, m = _parse_ym(close_month)
    hy, hm = _add_months(y, m, horizon_months)
    out = []
    for r in _msi_active_for_month(msi_rows, close_month):
        ey, em = r["end_year"], r["end_month"]
        if (y, m) <= (ey, em) <= (hy, hm):
            out.append(r)
    return out


def build(close_month: str, out_path: str) -> str:
    """Build the workbook and write to `out_path`. Returns the path."""
    transactions = load_transactions()
    recurring = load_recurring()

    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Budget"

    y, m = _parse_ym(close_month)
    row = 1

    # ---- Title ----
    ws.cell(row=row, column=1, value=f"MONTHLY BUDGET — {_month_label(y, m).upper()}").font = Font(
        bold=True, size=14
    )
    row += 1
    ws.cell(
        row=row, column=1,
        value=f"Amex cycle ends {dt.date(y, m, 19).strftime('%b %d, %Y')}",
    ).font = Font(italic=True, color="595959")
    row += 2  # blank row

    # ---- Header row ----
    ws.cell(row=row, column=1, value="Category")
    ws.cell(row=row, column=2, value="Amount (MXN)")
    ws.cell(row=row, column=3, value="% Income")
    ws.cell(row=row, column=4, value="Notes")
    for c in range(1, 5):
        ws.cell(row=row, column=c).fill = HEADER_FILL
        ws.cell(row=row, column=c).font = HEADER_FONT
    row += 1

    # ---- INCOME ----
    income_rows = recurring.get("income", []) or [
        {"label": "Net Monthly Income", "amount_mxn": 0}
    ]
    ws.cell(row=row, column=1, value="INCOME").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1
    income_start_row = row
    # We assume one net-income row for the pct formulas to reference (B6 in the
    # reference). If multiple income rows exist, we sum and place the total at
    # the first income row position.
    total_income = sum(float(r.get("amount_mxn", 0) or 0) for r in income_rows)
    primary_income_row = row
    label = income_rows[0].get("label", "Net Monthly Income")
    ws.cell(row=row, column=1, value=label)
    ws.cell(row=row, column=2, value=total_income).number_format = MXN_FMT
    ws.cell(row=row, column=4, value=income_rows[0].get("note", ""))
    row += 1
    row += 1  # blank

    pct_ref = f"B${primary_income_row}"

    def _pct_formula(data_row: int) -> str:
        return f"=B{data_row}/{pct_ref}"

    # ---- FIXED OBLIGATIONS ----
    fixed_rows = recurring.get("fixed_obligations", [])
    ws.cell(row=row, column=1, value="FIXED OBLIGATIONS").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1
    fixed_first = row
    for r in fixed_rows:
        ws.cell(row=row, column=1, value=r.get("label", ""))
        ws.cell(row=row, column=2, value=float(r.get("amount_mxn", 0) or 0)).number_format = MXN_FMT
        ws.cell(row=row, column=3, value=_pct_formula(row)).number_format = PCT_FMT
        due = r.get("due_day")
        if due:
            ws.cell(row=row, column=4, value=f"Due day {due}")
        row += 1
    fixed_last = row - 1
    ws.cell(row=row, column=1, value="SUBTOTAL FIXED").font = SUBTOTAL_FONT
    if fixed_first <= fixed_last:
        ws.cell(row=row, column=2, value=f"=SUM(B{fixed_first}:B{fixed_last})").number_format = MXN_FMT
    else:
        ws.cell(row=row, column=2, value=0).number_format = MXN_FMT
    ws.cell(row=row, column=3, value=_pct_formula(row)).number_format = PCT_FMT
    fixed_subtotal_row = row
    row += 2

    # ---- MSI ----
    msi_active = _msi_active_for_month(recurring.get("msi", []), close_month)
    ws.cell(row=row, column=1, value="MSI INSTALLMENT PLANS").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1
    msi_first = row
    for r in msi_active:
        cur = r.get("current_installment", "?")
        total = r.get("months_total", "?")
        ending = (r.get("end_year"), r.get("end_month")) == (y, m)
        label = f"{r.get('label','')} ({cur} of {total}" + (" — DONE)" if ending else ")")
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=float(r.get("amount_mxn", 0) or 0)).number_format = MXN_FMT
        ws.cell(row=row, column=3, value=_pct_formula(row)).number_format = PCT_FMT
        if ending:
            ws.cell(row=row, column=4, value="Last payment this cycle")
        row += 1
    msi_last = row - 1
    ws.cell(row=row, column=1, value="SUBTOTAL MSI").font = SUBTOTAL_FONT
    if msi_first <= msi_last:
        ws.cell(row=row, column=2, value=f"=SUM(B{msi_first}:B{msi_last})").number_format = MXN_FMT
    else:
        ws.cell(row=row, column=2, value=0).number_format = MXN_FMT
    ws.cell(row=row, column=3, value=_pct_formula(row)).number_format = PCT_FMT
    msi_subtotal_row = row
    row += 2

    # ---- SUBSCRIPTIONS ----
    subs = [s for s in recurring.get("subscriptions", []) if s.get("active", True)]
    ws.cell(row=row, column=1, value="SUBSCRIPTIONS & RECURRING").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1
    subs_first = row
    for s in subs:
        ws.cell(row=row, column=1, value=s.get("label", ""))
        ws.cell(row=row, column=2, value=float(s.get("amount_mxn", 0) or 0)).number_format = MXN_FMT
        ws.cell(row=row, column=3, value=_pct_formula(row)).number_format = PCT_FMT
        row += 1
    subs_last = row - 1
    ws.cell(row=row, column=1, value="SUBTOTAL SUBSCRIPTIONS").font = SUBTOTAL_FONT
    if subs_first <= subs_last:
        ws.cell(row=row, column=2, value=f"=SUM(B{subs_first}:B{subs_last})").number_format = MXN_FMT
    else:
        ws.cell(row=row, column=2, value=0).number_format = MXN_FMT
    ws.cell(row=row, column=3, value=_pct_formula(row)).number_format = PCT_FMT
    subs_subtotal_row = row
    row += 2

    # ---- OFF-CARD ----
    off_card = recurring.get("off_card", [])
    ws.cell(row=row, column=1, value="OFF-CARD EXPENSES (BBVA / Cash)").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1
    off_first = row
    for r in off_card:
        ws.cell(row=row, column=1, value=r.get("label", ""))
        ws.cell(row=row, column=2, value=float(r.get("amount_mxn", 0) or 0)).number_format = MXN_FMT
        ws.cell(row=row, column=3, value=_pct_formula(row)).number_format = PCT_FMT
        row += 1
    off_last = row - 1
    ws.cell(row=row, column=1, value="SUBTOTAL OFF-CARD").font = SUBTOTAL_FONT
    if off_first <= off_last:
        ws.cell(row=row, column=2, value=f"=SUM(B{off_first}:B{off_last})").number_format = MXN_FMT
    else:
        ws.cell(row=row, column=2, value=0).number_format = MXN_FMT
    ws.cell(row=row, column=3, value=_pct_formula(row)).number_format = PCT_FMT
    off_subtotal_row = row
    row += 2

    # ---- VARIABLE SPENDING (from transactions) ----
    var_totals = _sum_by_subcategory(transactions, "VARIABLE", close_month)
    ws.cell(row=row, column=1, value="VARIABLE SPENDING (Personal)").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1
    var_first = row
    for sub, amt in sorted(var_totals.items(), key=lambda kv: -kv[1]):
        ws.cell(row=row, column=1, value=sub)
        ws.cell(row=row, column=2, value=amt).number_format = MXN_FMT
        ws.cell(row=row, column=3, value=_pct_formula(row)).number_format = PCT_FMT
        row += 1
    var_last = row - 1
    ws.cell(row=row, column=1, value="SUBTOTAL VARIABLE").font = SUBTOTAL_FONT
    if var_first <= var_last:
        ws.cell(row=row, column=2, value=f"=SUM(B{var_first}:B{var_last})").number_format = MXN_FMT
    else:
        ws.cell(row=row, column=2, value=0).number_format = MXN_FMT
    ws.cell(row=row, column=3, value=_pct_formula(row)).number_format = PCT_FMT
    var_subtotal_row = row
    row += 2

    # ---- BUSINESS REIMBURSABLE ----
    biz_rows = [
        t for t in transactions
        if t.get("close_month") == close_month and t.get("section") == "BUSINESS_REIMBURSABLE"
    ]
    ws.cell(row=row, column=1, value="BUSINESS EXPENSES (Reimbursable)").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1
    biz_first = row
    for t in sorted(biz_rows, key=lambda t: t.get("date", "")):
        amt = t.get("amount_mxn") or t.get("amount_native") or 0
        reimb = t.get("reimbursement") or {}
        label = t.get("description", "")
        date_s = t.get("date", "")
        reem = reimb.get("reembolso", "")
        note_bits = [date_s]
        if reem:
            note_bits.append(reem)
        ws.cell(row=row, column=1, value=f"{label} ({date_s})")
        ws.cell(row=row, column=2, value=float(amt)).number_format = MXN_FMT
        ws.cell(row=row, column=4, value=" | ".join(note_bits))
        row += 1
    biz_last = row - 1
    ws.cell(row=row, column=1, value="TOTAL REIMBURSABLE").font = SUBTOTAL_FONT
    if biz_first <= biz_last:
        ws.cell(row=row, column=2, value=f"=SUM(B{biz_first}:B{biz_last})").number_format = MXN_FMT
    else:
        ws.cell(row=row, column=2, value=0).number_format = MXN_FMT
    ws.cell(row=row, column=4, value="Submit ASAP")
    biz_total_row = row
    row += 2

    # ---- FEES ----
    fee_totals = _sum_by_subcategory(transactions, "FEES", close_month)
    ws.cell(row=row, column=1, value="FEES & PENALTIES").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1
    fee_first = row
    for sub, amt in sorted(fee_totals.items(), key=lambda kv: -kv[1]):
        ws.cell(row=row, column=1, value=sub)
        ws.cell(row=row, column=2, value=amt).number_format = MXN_FMT
        row += 1
    fee_last = row - 1
    ws.cell(row=row, column=1, value="SUBTOTAL FEES").font = SUBTOTAL_FONT
    if fee_first <= fee_last:
        ws.cell(row=row, column=2, value=f"=SUM(B{fee_first}:B{fee_last})").number_format = MXN_FMT
    else:
        ws.cell(row=row, column=2, value=0).number_format = MXN_FMT
    ws.cell(row=row, column=3, value=_pct_formula(row)).number_format = PCT_FMT
    ws.cell(row=row, column=4, value="Avoid next month")
    fee_subtotal_row = row
    row += 2

    # ---- SUMMARY ----
    ws.cell(row=row, column=1, value="SUMMARY").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1
    ws.cell(row=row, column=1, value="Monthly Income")
    ws.cell(row=row, column=2, value=f"=B{primary_income_row}").number_format = MXN_FMT
    row += 1
    ws.cell(row=row, column=1, value="Total Recurring (Fixed + MSI + Subs + Off-Card)")
    ws.cell(
        row=row, column=2,
        value=f"=B{fixed_subtotal_row}+B{msi_subtotal_row}+B{subs_subtotal_row}+B{off_subtotal_row}",
    ).number_format = MXN_FMT
    ws.cell(row=row, column=3, value=_pct_formula(row)).number_format = PCT_FMT
    recurring_row = row
    row += 1
    ws.cell(row=row, column=1, value="Variable Personal Spending")
    ws.cell(row=row, column=2, value=f"=B{var_subtotal_row}").number_format = MXN_FMT
    ws.cell(row=row, column=3, value=_pct_formula(row)).number_format = PCT_FMT
    variable_row = row
    row += 1
    ws.cell(row=row, column=1, value="Fees & Penalties")
    ws.cell(row=row, column=2, value=f"=B{fee_subtotal_row}").number_format = MXN_FMT
    ws.cell(row=row, column=3, value=_pct_formula(row)).number_format = PCT_FMT
    fees_row = row
    row += 1
    ws.cell(row=row, column=1, value="TOTAL PERSONAL SPEND").font = SUBTOTAL_FONT
    ws.cell(
        row=row, column=2,
        value=f"=B{recurring_row}+B{variable_row}+B{fees_row}",
    ).number_format = MXN_FMT
    ws.cell(row=row, column=3, value=_pct_formula(row)).number_format = PCT_FMT
    total_spend_row = row
    row += 2

    ws.cell(row=row, column=1, value="MONTHLY SURPLUS (Income − Personal Spend)").font = SUBTOTAL_FONT
    ws.cell(
        row=row, column=2,
        value=f"=B{primary_income_row}-B{total_spend_row}",
    ).number_format = MXN_FMT
    ws.cell(row=row, column=3, value=_pct_formula(row)).number_format = PCT_FMT
    row += 2

    ws.cell(row=row, column=1, value="Pending Business Reimbursement")
    ws.cell(row=row, column=2, value=f"=B{biz_total_row}").number_format = MXN_FMT
    ws.cell(row=row, column=4, value="Cash coming back to you")
    row += 2

    # ---- MSI ROLL-OFF TRACKER ----
    ending_soon = _msi_ending_soon(recurring.get("msi", []), close_month, horizon_months=3)
    ws.cell(row=row, column=1, value="MSI ROLL-OFF TRACKER").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1
    rolloff_first = row
    for r in ending_soon:
        cur = r.get("current_installment", "?")
        total = r.get("months_total", "?")
        ending_month = f"{r.get('end_year'):04d}-{r.get('end_month'):02d}"
        ws.cell(row=row, column=1, value=f"{r.get('label','')} ({cur}/{total}) — ends {ending_month}")
        ws.cell(row=row, column=2, value=float(r.get("amount_mxn", 0) or 0)).number_format = MXN_FMT
        ws.cell(row=row, column=4, value="Frees up when it ends")
        row += 1
    rolloff_last = row - 1
    if rolloff_first <= rolloff_last:
        ws.cell(row=row, column=1, value=f"Total freed by end of horizon").font = SUBTOTAL_FONT
        ws.cell(
            row=row, column=2,
            value=f"=SUM(B{rolloff_first}:B{rolloff_last})",
        ).number_format = MXN_FMT
    else:
        ws.cell(row=row, column=1, value="(no MSI ending in next 3 months)").font = Font(italic=True)

    # ---- Column widths ----
    widths = {1: 50, 2: 16, 3: 12, 4: 50}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)
    return out_path
